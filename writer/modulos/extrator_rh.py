# modulos/extrator_rh.py
# -*- coding: utf-8 -*-
"""
Extrator da 'Relação de Recursos Humanos' para dois modelos de formulário:
- MODELO NOVO (tabelado): "Item CPF NOME TITULAÇÃO FUNÇÃO TOTAL HORAS (ANUAL) DEDICAÇÃO VALOR R$"
  Abordagem: Layout (PyMuPDF "words") ancorada no CPF + cluster (linha atual + próxima).
- MODELO ANTIGO (rotulado): linhas do tipo "Item N", "CPF", "Nome", "Titulação",
  "Total Horas (Anual)", "Dedicação", "Valor (R$)".
  Abordagem: Texto contínuo (PyMuPDF "text") + regex de rótulos, tolerante a quebras de página.
API principal:
 extract_rh_entries(source: bytes\bytearray\str, is_text=False, debug=False) -> RHExtractResult
"""
from __future__ import annotations
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple, Union
import re
import unicodedata

# ========================= Utils =========================
def _norm_spaces(s: str) -> str:
    if s is None: return ""
    # normaliza whitespaces e remove marcadores de bold (**)
    s = s.replace("\u00A0", " ")
    s = s.replace("\r", "\n").replace("\t", " ")
    s = s.replace("*", "")
    s = re.sub(r"[\f\v]+", " ", s)
    s = re.sub(r"[\t]+\n", "\n", s)
    return s

def _strip_accents_lower(s: str) -> str:
    s = (s or "").replace("\u00A0", " ")
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c)).lower()

def _parse_int_pt(s: str) -> Optional[int]:
    if not s: return None
    digits = re.sub(r"[^\d]", "", s)
    try: return int(digits)
    except: return None

def _parse_float_br(s: str) -> Optional[float]:
    if not s: return None
    keep = re.sub(r"[^0-9,\.]", "", s.strip())
    if not keep: return None
    keep = keep.replace(".", "").replace(",", ".")
    try: return float(keep)
    except: return None

def _parse_cpf_digits(s: str) -> Optional[str]:
    if not s: return None
    digits = re.sub(r"\D+", "", s)
    return digits if digits else None

def _fmt_cpf(d: Optional[str]) -> Optional[str]:
    if d and len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    return None

def _extract_hours(text: str) -> Optional[int]:
    """
    Extrai horas aceitando:
    - milhar com espaço: "2 884"
    - milhar com ponto: "1.089"
    - inteiros simples: "731"
    Regra: pega o ÚLTIMO número plausível do trecho analisado (antes da Dedicação/Valor).
    """
    if not text: return None
    # milhar com espaço (um ou mais grupos 3)
    mg = re.findall(r"\b\d{1,3}(?:\s\d{3})+\b", text)
    if mg:
        return _parse_int_pt(re.sub(r"\s+", "", mg[-1]))
    # milhar com ponto, ou simples
    mg = re.findall(r"\b\d{1,4}(?:\.\d{3})*\b", text)
    if mg:
        return _parse_int_pt(mg[-1])
    return None

# ========================= Regex base =========================
HIFEN = r"[\-\u2010-\u2015]"
CPF_FMT_RX = re.compile(rf"\b\d{{3}}\.\d{{3}}\.\d{{3}}{HIFEN}\d{{2}}\b")
CPF_DIGITS_RX = re.compile(r"\b\d{11}\b")

# Titulações mais comuns (masc/fem, com/sem acentos)
TITULACOES = [
    "Técnico de Nível Médio", "Tecnico de Nivel Medio",
    "Técnica de Nível Médio", "Tecnica de Nivel Medio",
    "Graduado", "Graduada",
    "Pós-Grad[uú]ado", "Pos-Grad[uú]ado", "P[oó]s Grad[uú]ado",
    "Pós-Grad[uú]ada", "Pos-Grad[uú]ada", "P[oó]s Grad[uú]ada",
    "Tecnólogo", "Tecnologa", "Tecnologo",
    "Mestre", "Doutor", "Doutora",
    "Apoio Técnico", "Apoio Tecnico",
]
TIT_RX = re.compile(r"(?i)\b(" + "|".join(t.replace(" ", r"\s+") for t in TITULACOES) + r")\b")
DEDIC_RX = re.compile(r"(?i)\b(Parcial|Exclusiva)\b")
VALOR_ANY_RX = re.compile(r"R\$\s*([\d\.\s]+,\d{2})", re.I)

# Âncoras de seção (antigo)
ANCHOR_RH_RX = re.compile(r"(?i)rela[cç][aã]o\s+de\s+recursos\s+humanos")
HEADER_INLINE_RX = re.compile(r"(?i)\bItem\s+CPF\s+NOME\s+TITULA[cç][aã]o\s+FUN[cç][aã]o\s+TOTAL\s+HORAS")
TOTAL_RS_ANY_RX = re.compile(r"(?i)Total\s*R\$\s*[\d\.\s]+,\d{2}")
NEXT_HEADING_RX = re.compile(r"(?mi)^\s*(DISP[EÊ]NDIOS|FONTES DE FINANCIAMENTO|GASTOS DESTINADOS)\b")
DESCR_HEADING_RX = re.compile(r"(?mi)^\s*DESCREVA\s+AS\s+ATIVIDADES\b")

# Rótulos (antigo)
RX_ITEM = re.compile(r"(?i)^Item\s+(\d{1,4})\b")
RX_CPF  = re.compile(r"^CPF\b\s*[:\-\–—]?\s*(.+)$", re.I)
RX_NOME = re.compile(r"^Nome\b\s*[:\-\–—]?\s*(.+)$", re.I)
RX_TIT  = re.compile(r"^Titula[cç][aã]o\b\s*[:\-\–—]?\s*(.+)$", re.I)
RX_TOT  = re.compile(r"^Total\s+Horas(?:\s*\(\s*Anual\s*\))?\b.*?(\d[\d\.\s]*)\s*$", re.I)
RX_DED  = re.compile(r"^Dedica[cç][aã]o\b\s*[:\-\–—]?\s*(.+)$", re.I)
RX_VAL  = re.compile(r"^Valor(?:\s*\(\s*R\$\s*\))?\b.*?([R$ ]?[\d\.\s]+(?:,\d{2})?)\s*$", re.I)

# ========================= Retorno =========================
@dataclass
class RHExtractResult:
    items: List[Dict[str, Optional[Union[str, int, float]]]]
    total_text: Optional[str]
    section_bounds: Tuple[int, int]
    model: str  # "novo" | "antigo" | "desconhecido"
    debug_samples: Optional[List[Dict]] = None

# ========================= PyMuPDF =========================
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

# ========================= Layout helpers (novo) =========================
def _lines_from_words(words, y_tol: float = 3.5):
    """Agrupa words em linhas por proximidade de y."""
    if not words: return []
    words_sorted = sorted(words, key=lambda w: (w[1], w[0]))  # y0, x0
    lines, cur, cur_y = [], [], None
    for (x0,y0,x1,y1,t) in words_sorted:
        if cur_y is None or abs(y0-cur_y) <= y_tol:
            cur.append((x0,y0,x1,y1,t)); cur_y = y0 if cur_y is None else cur_y
        else:
            lines.append((cur_y, sorted(cur, key=lambda z: z[0])))
            cur, cur_y = [(x0,y0,x1,y1,t)], y0
    if cur:
        lines.append((cur_y, sorted(cur, key=lambda z: z[0])))
    return lines

def _join_line(line_words) -> str:
    return " ".join(w[4] for w in line_words).strip()

def _find_cpfes_in_line(line_words) -> List[Tuple[int,int,str]]:
    """Retorna [(idx_word, len_words, cpf_digits)] por CPF encontrado."""
    texts = [w[4] for w in line_words]
    res = []
    for i, w in enumerate(texts):
        if CPF_FMT_RX.fullmatch(w):
            cd = _parse_cpf_digits(w)
            if cd and len(cd)==11:
                res.append((i,1,cd))
    if not res:
        # CPF sem pontuação (conservador)
        joined = " ".join(texts)
        m = CPF_DIGITS_RX.search(joined)
        if m:
            cd = m.group(0)
            res.append((0,1,cd))
    return res

def _collect_cluster(lines, li, two_lines=True):
    """Linha atual + próxima linha (ajuda quando NOME/TITULAÇÃO quebram)."""
    _, line_words = lines[li]
    cluster = list(line_words)
    if two_lines and li+1 < len(lines):
        _, line2 = lines[li+1]
        cluster += line2
    return sorted(cluster, key=lambda z: (z[1], z[0]))

def _parse_cluster(cluster_words, cpf_digits, debug=False) -> Tuple[Optional[Dict], Optional[Dict]]:
    """Extrai campos do cluster (1–2 linhas) ancorado no CPF."""
    toks = sorted(cluster_words, key=lambda z: z[0])  # por x
    texts = [t[4] for t in toks]
    # localiza o índice do token do CPF
    cpf_idx = None
    for i, tx in enumerate(texts):
        if (CPF_FMT_RX.fullmatch(tx) or CPF_DIGITS_RX.fullmatch(tx)) and _parse_cpf_digits(tx) == cpf_digits:
            cpf_idx = i; break
    # Item = último inteiro à ESQUERDA do CPF
    item_num = None
    if cpf_idx is not None:
        x_cpf = toks[cpf_idx][0]
        left_ints = [(t[0], re.findall(r"\b\d{1,4}\b", tx)) for (t,tx) in zip(toks, texts) if t[0] < x_cpf]
        if left_ints:
            flat = [(x, d) for (x, arr) in left_ints for d in arr]
            if flat:
                item_num = _parse_int_pt(flat[-1][1])
    # texto após CPF
    after = " ".join(texts[cpf_idx+1:] if cpf_idx is not None else texts)
    # Titulação
    nome, titulacao, funcao, dedicacao, valor, horas = None, None, None, None, None, None
    mt = TIT_RX.search(after)
    if mt:
        nome = after[:mt.start()].strip()
        titulacao = mt.group(0).strip()
        tail = after[mt.end():].strip()
    else:
        nome = after.strip()
        tail = ""
    # Valor (último R$ no cluster inteiro)
    cluster_text = " ".join(texts)
    mv_all = list(VALOR_ANY_RX.finditer(cluster_text))
    if mv_all:
        valor = _parse_float_br(mv_all[-1].group(1))
    # Dedicação
    tgt_for_ded = tail if mt else after
    md = DEDIC_RX.search(tgt_for_ded)
    if md:
        dedicacao = md.group(1).capitalize()
    # Horas (antes da dedicação, se houver; senão no tail todo)
    search_zone = tgt_for_ded
    if md:
        search_zone = tgt_for_ded[:md.start()].strip()
    horas = _extract_hours(search_zone)
    if horas is not None:
        # Função = tudo antes da expressão de horas
        hs = str(horas)
        # casa tanto "2884" quanto "2 884" e "2.884"
        m_alt = re.search(rf"\b{hs[:1]}[ \.]?{hs[1:]}\b", search_zone)
        pos = m_alt.start() if m_alt else search_zone.rfind(hs)
        funcao = search_zone[:pos].strip() if pos != -1 else (search_zone or None)
    else:
        funcao = search_zone or None
    # Limpeza de múltiplos espaços
    def _clean(s: Optional[str]) -> Optional[str]:
        if not s: return s
        return re.sub(r"\s{2,}", " ", s).strip()
    nome = _clean(nome)
    funcao = _clean(funcao)
    rec = {
        "item_numero": item_num,
        "cpf": cpf_digits,
        "cpf_formatado": _fmt_cpf(cpf_digits),
        "nome": nome or None,
        "titulacao": titulacao or None,
        "funcao": funcao or None,
        "total_horas_anual": horas,
        "dedicacao": dedicacao,
        "valor_rs": valor,
        "raw": cluster_text,
    }
    debug_blob = None
    if debug:
        debug_blob = {"cpf": cpf_digits, "cluster_text": cluster_text, "after": after, "parsed": rec}
    if rec["cpf"] and rec["item_numero"] is not None:
        return rec, debug_blob
    return None, debug_blob

# ========================= Extração MODELO NOVO =========================
def _extract_by_layout_cpf(raw_pdf: Union[bytes, bytearray], debug=False) -> Tuple[List[Dict], Optional[str], List[Dict]]:
    if fitz is None:
        raise RuntimeError("PyMuPDF (fitz) não está instalado.")
    items: List[Dict] = []
    dbg_list: List[Dict] = []
    total_text: Optional[str] = None
    with fitz.open(stream=raw_pdf, filetype="pdf") as doc:
        for page in doc:
            words = [(w[0], w[1], w[2], w[3], w[4]) for w in page.get_text("words")]
            lines = _lines_from_words(words, y_tol=3.5)
            i = 0
            while i < len(lines):
                _, line_words = lines[i]
                line_txt = _join_line(line_words)
                # total R$ (às vezes vem junto do RH ainda)
                m_total = TOTAL_RS_ANY_RX.search(line_txt)
                if m_total:
                    total_text = m_total.group(0).strip()
                # CPFs na linha
                cpfs = _find_cpfes_in_line(line_words)
                if cpfs:
                    cluster = _collect_cluster(lines, i, two_lines=True)
                    # pode haver mais de um CPF na mesma linha (caso raro). Trata todos.
                    for (_, _, cpf_digits) in cpfs:
                        rec, db = _parse_cluster(cluster, cpf_digits, debug=debug)
                        if db: dbg_list.append(db)
                        if rec: items.append(rec)
                i += 1
    return items, total_text, dbg_list

# ========================= Extração MODELO ANTIGO =========================
def _extract_text_from_pdf(raw_pdf: Union[bytes, bytearray]) -> str:
    if fitz is None:
        raise RuntimeError("PyMuPDF (fitz) não está instalado.")
    parts: List[str] = []
    with fitz.open(stream=raw_pdf, filetype="pdf") as doc:
        for p in doc:
            parts.append(p.get_text("text"))
    return _norm_spaces("\n".join(parts))

def _extract_section_by_text(full_text: str) -> Tuple[str, int, int, Optional[str]]:
    if not full_text: return "", -1, -1, None
    raw = _norm_spaces(full_text)
    m_start = ANCHOR_RH_RX.search(raw) or HEADER_INLINE_RX.search(raw)
    if not m_start: return "", -1, -1, None
    start = m_start.end()
    after = raw[start:]
    # delimita até a próxima seção ou até o "DESCREVA..." (quando o modelo novo adiciona narrativas)
    ends = [len(after)]
    for rx in (TOTAL_RS_ANY_RX, NEXT_HEADING_RX, DESCR_HEADING_RX):
        m = rx.search(after)
        if m: ends.append(m.start())
    end_rel = min(ends)
    section = after[:end_rel]
    total_line = None
    m_amt = TOTAL_RS_ANY_RX.search(after[: end_rel+200])
    if m_amt: total_line = m_amt.group(0).strip()
    return section, start, start+end_rel, total_line

def _parse_modelo_antigo(section_text: str) -> List[Dict]:
    # força quebra antes de rótulos e "Item N"
    s = re.sub(r"(?i)\b(Item\s+\d{1,4})\b", r"\n\1", section_text)
    s = re.sub(r"(?i)\b(CPF|Nome|Titula[cç][aã]o|Total\s+Horas\s*\(\s*Anual\s*\)|Dedica[cç][aã]o|Valor\s*\(\s*R\$\s*\))\b", r"\n\1", s)
    s = re.sub(r"[ \t]+\n", "\n", s)
    s = re.sub(r"\n{2,}", "\n", s).strip()
    items: List[Dict] = []
    cur: Optional[Dict] = None; raw_lines: List[str] = []
    for line in s.split("\n"):
        t = (line or "").strip()
        if not t: continue
        m_item = RX_ITEM.match(t)
        if m_item:
            if cur is not None:
                cur["raw"] = "\n".join(raw_lines).strip(); items.append(cur)
            raw_lines = [t]
            cur = {"item_numero": int(m_item.group(1)), "cpf":None, "cpf_formatado":None, "nome":None,
                   "titulacao":None, "funcao":None, "total_horas_anual":None, "dedicacao":None, "valor_rs":None, "raw":""}
            continue
        if cur is None: continue
        raw_lines.append(t)
        if (m := RX_CPF.match(t)):
            d = _parse_cpf_digits(m.group(1)); cur["cpf"]=d; cur["cpf_formatado"]=_fmt_cpf(d); continue
        if (m := RX_NOME.match(t)): cur["nome"]=m.group(1).strip(); continue
        if (m := RX_TIT.match(t)):  cur["titulacao"]=m.group(1).strip(); continue
        if (m := RX_TOT.match(t)):  cur["total_horas_anual"]=_parse_int_pt(m.group(1)); continue
        if (m := RX_DED.match(t)):  cur["dedicacao"]=m.group(1).strip(); continue
        if (m := RX_VAL.match(t)):  cur["valor_rs"]=_parse_float_br(m.group(1)); continue
    if cur is not None:
        cur["raw"] = "\n".join(raw_lines).strip(); items.append(cur)
    return items

# ========================= Detector de modelo =========================
def _sniff_model(full_text: str) -> str:
    text = _norm_spaces(full_text)
    has_table_header = bool(HEADER_INLINE_RX.search(text))
    has_rh_anchor = bool(ANCHOR_RH_RX.search(text))
    if has_table_header:
        return "novo"
    if has_rh_anchor:
        # se não há cabeçalho de tabela, assume rótulos
        return "antigo"
    return "desconhecido"

# ========================= Função Pública =========================
def extract_rh_entries(
    source: Union[bytes, bytearray, str],
    *,
    is_text: bool = False,
    capture_total_line: bool = True,
    debug: bool = False,
) -> RHExtractResult:
    """
    - source bytes/bytearray (PDF bytes) OU str (texto plain do PDF)
    - is_text: True quando 'source' já é texto.
    - debug: inclui amostras de clusters (modelo novo).
    """
    if is_text:
        full_text = _norm_spaces(str(source))
        model = _sniff_model(full_text)
        if model == "novo":
            # sem PyMuPDF não dá para fazer layout; retorna básico
            return RHExtractResult(items=[], total_text=None, section_bounds=(-1,-1), model=model)
        # modelo antigo por texto
        section_text, i0, i1, total_line = _extract_section_by_text(full_text)
        items = _parse_modelo_antigo(section_text) if i0 != -1 else []
        return RHExtractResult(
            items=items, total_text=(total_line if capture_total_line else None),
            section_bounds=(i0, i1), model=("antigo" if i0 != -1 else "desconhecido")
        )

    # --- quando recebemos bytes de PDF ---
    raw_pdf = source  # type: ignore

    # 1) sniff do modelo a partir de texto leve do PDF
    try:
        full_text = _extract_text_from_pdf(raw_pdf)
    except Exception:
        full_text = ""
    model = _sniff_model(full_text)

    # 2) MODELO NOVO (tabela) → layout CPF-ancorado
    if model == "novo" and isinstance(raw_pdf, (bytes, bytearray)):
        try:
            items, total, dbg = _extract_by_layout_cpf(raw_pdf, debug=debug)
            if items:
                return RHExtractResult(
                    items=items, total_text=(total if capture_total_line else None),
                    section_bounds=(-1,-1), model="novo",
                    debug_samples=(dbg[:25] if debug else None)
                )
        except Exception:
            # fallback para texto (mínimo)
            pass

    # 3) MODELO ANTIGO (rótulos) → texto por seção
    section_text, i0, i1, total_line = _extract_section_by_text(full_text)
    items_a = _parse_modelo_antigo(section_text) if i0 != -1 else []

    # Se não retornou nada e suspeita de ser "novo", tenta mesmo assim a via layout
    if not items_a and isinstance(raw_pdf, (bytes, bytearray)):
        try:
            items_b, total_b, dbg = _extract_by_layout_cpf(raw_pdf, debug=debug)
            if items_b:
                return RHExtractResult(
                    items=items_b, total_text=(total_b if capture_total_line else None),
                    section_bounds=(-1,-1), model=("novo" if model != "antigo" else model),
                    debug_samples=(dbg[:25] if debug else None)
                )
        except Exception:
            pass

    return RHExtractResult(
        items=items_a,
        total_text=(total_line if capture_total_line else None),
        section_bounds=(i0, i1),
        model=("antigo" if items_a else model),
    )

# ========================= Saída: Planilha Modelo (sem template externo) =========================
# Este bloco cria um .xlsx com a aba e cabeçalhos do arquivo:
# "Contestação RH modelo_v3_220425.xlsm" (Planilha1) — sem depender do próprio .xlsm.
# Colunas: Item | CPF | Nome | Titulação | Total de horas | Dedicação |
#          Total Declarado (R$) | Total Glosado (R$) | Justificativa
# Observações:
# - "Justificativa" existirá e ficará em branco, como solicitado.
# - "Total Glosado (R$)" AGORA repete o mesmo valor de "Total Declarado (R$)".
# - "CPF" usa "cpf_formatado" quando houver, caindo para "cpf" (dígitos).
# - "Total Declarado (R$)" recebe "valor_rs"; "Total de horas" recebe "total_horas_anual".
try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # evitar falha de import para quem só usa o extrator

def export_rh_to_model_sheet(
    result: "RHExtractResult",
    output_path: str,
    *,
    sheet_name: str = "Planilha1",
    autofit: bool = True,
    write_total_text_cell: str = None,  # ex.: "J2" (opcional)
) -> str:
    """
    Gera um .xlsx novo com o layout do seu modelo (sem necessidade de template externo).

    Parâmetros:
        result: RHExtractResult de extract_rh_entries(...)
        output_path: caminho do arquivo .xlsx a ser gerado
        sheet_name: nome da aba (default: "Planilha1")
        autofit: ajusta largura básica das colunas (heurística simples)
        write_total_text_cell: se informado, escreve result.total_text nessa célula (texto "Total R$ ...")

    Retorna:
        Caminho salvo (output_path).
    """
    if Workbook is None:
        raise RuntimeError("openpyxl não está instalado. Instale com: pip install openpyxl")

    headers = [
        "Item",
        "CPF",
        "Nome",
        "Titulação",
        "Total de horas",
        "Dedicação",
        "Total Declarado (R$)",
        "Total Glosado (R$)",
        "Justificativa",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Cabeçalhos
    ws.append(headers)

    # Linhas
    for it in (result.items or []):
        row = [
            it.get("item_numero"),
            it.get("cpf_formatado") or it.get("cpf"),
            it.get("nome"),
            it.get("titulacao"),
            it.get("total_horas_anual"),
            it.get("dedicacao"),
            it.get("valor_rs"),  # Total Declarado (R$)
            it.get("valor_rs"),  # Total Glosado (R$) → repete o mesmo valor, conforme solicitado
            None,                # Justificativa
        ]
        ws.append(row)

    # (Opcional) texto do total em uma célula específica
    if write_total_text_cell and result.total_text:
        ws[write_total_text_cell] = result.total_text

    # Ajuste simples de largura
    if autofit:
        from openpyxl.utils import get_column_letter
        for col_idx, header in enumerate(headers, start=1):
            maxlen = len(str(header))
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is not None:
                    maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, maxlen + 2), 60)

    wb.save(output_path)
